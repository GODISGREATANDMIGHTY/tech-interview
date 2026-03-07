
Category 1: Regex-Based Exercises (35)

import re

# 1. Extract all email addresses from a string
text = "Contact us at info@example.com or support@site.org"
emails = re.findall(r'\b[\w.-]+@[\w.-]+\.\w+\b', text)
print(emails)

# 2. Validate a list of phone numbers (simple US format)
phones = ["123-456-7890", "555-999-8888", "1234567"]
valid_phones = [p for p in phones if re.fullmatch(r'\d{3}-\d{3}-\d{4}', p)]
print(valid_phones)

# 3. Find all URLs in a text file
text = "Visit https://example.com and http://site.org"
urls = re.findall(r'https?://[^\s]+', text)
print(urls)

# 4. Replace multiple spaces with a single space
s = "This   is   spaced"
cleaned = re.sub(r'\s+', ' ', s)
print(cleaned)

# 5. Remove all HTML tags from a string
html = "<p>Hello <b>World</b></p>"
text_only = re.sub(r'<.*?>', '', html)
print(text_only)

# 6. Extract hashtags from tweets
tweet = "Loving #Python and #AI"
hashtags = re.findall(r'#\w+', tweet)
print(hashtags)

# 7. Detect valid IP addresses
ips = ["192.168.0.1", "256.0.0.1"]
valid_ips = [ip for ip in ips if re.fullmatch(r'(?:\d{1,3}\.){3}\d{1,3}', ip)]
print(valid_ips)

# 8. Split a text by punctuation marks
text = "Hello, world! How are you?"
words = re.split(r'[,.!?]\s*', text)
print(words)

# 9. Validate date strings (YYYY-MM-DD)
dates = ["2023-03-07", "07-03-2023"]
valid_dates = [d for d in dates if re.fullmatch(r'\d{4}-\d{2}-\d{2}', d)]
print(valid_dates)

# 10. Find all words starting with a capital letter
text = "Hello World from Python"
caps = re.findall(r'\b[A-Z][a-z]*\b', text)
print(caps)

# 11. Extract all numeric values from a string
text = "There are 12 apples and 30 oranges"
nums = re.findall(r'\d+', text)
print(nums)

# 12. Detect repeated words in a text
text = "This is is repeated"
repeats = re.findall(r'\b(\w+)\s+\1\b', text)
print(repeats)

# 13. Replace all digits with #
s = "My phone is 123-456-7890"
masked = re.sub(r'\d', '#', s)
print(masked)

# 14. Extract domain names from URLs
urls = ["https://example.com/path", "http://site.org"]
domains = [re.search(r'https?://([^/]+)', u).group(1) for u in urls]
print(domains)

# 15. Validate hexadecimal color codes
colors = ["#FFF", "#123ABC", "123456"]
valid_colors = [c for c in colors if re.fullmatch(r'#(?:[0-9A-Fa-f]{3}){1,2}', c)]
print(valid_colors)

# 16. Remove all non-alphanumeric characters
s = "Hello, World! 123"
cleaned = re.sub(r'[^A-Za-z0-9 ]+', '', s)
print(cleaned)

# 17. Detect time in HH:MM format
times = ["12:30", "24:60", "09:15"]
valid_times = [t for t in times if re.fullmatch(r'([01]\d|2[0-3]):[0-5]\d', t)]
print(valid_times)

# 18. Extract all mentions (e.g., @username)
tweet = "Hello @user1 and @user2"
mentions = re.findall(r'@\w+', tweet)
print(mentions)

# 19. Find words that end with a specific suffix
text = "I like running, swimming, coding"
suffix_words = re.findall(r'\b\w+ing\b', text)
print(suffix_words)

# 20. Validate US ZIP codes
zips = ["12345", "12345-6789", "ABCDE"]
valid_zips = [z for z in zips if re.fullmatch(r'\d{5}(-\d{4})?', z)]
print(valid_zips)

# 21. Replace dates in text with ISO format
text = "Today is 07/03/2026"
converted = re.sub(r'(\d{2})/(\d{2})/(\d{4})', r'\3-\2-\1', text)
print(converted)

# 22. Extract prices from receipts
text = "Item A $12.50, Item B $7.99"
prices = re.findall(r'\$\d+\.\d{2}', text)
print(prices)

# 23. Check for valid credit card numbers (simple pattern)
cards = ["1234-5678-9012-3456", "1234 5678 9012 3456"]
valid_cards = [c for c in cards if re.fullmatch(r'(\d{4}[- ]?){3}\d{4}', c)]
print(valid_cards)

# 24. Extract filenames from file paths
paths = ["/home/user/file.txt", "C:\\folder\\data.csv"]
filenames = [re.search(r'[^\\/]+$', p).group(0) for p in paths]
print(filenames)

# 25. Replace multiple newlines with a single newline
text = "Line1\n\n\nLine2"
single_newline = re.sub(r'\n+', '\n', text)
print(single_newline)

# 26. Find all words of a certain length
text = "I love Python programming"
words = re.findall(r'\b\w{6}\b', text)
print(words)

# 27. Extract all JSON-like key-value patterns
text = '{"name": "John", "age": 30}'
kv_pairs = re.findall(r'"(\w+)":\s*"?(.*?)"?[,}]', text)
print(kv_pairs)

# 28. Detect duplicate lines in a text
text = "Hello\nWorld\nHello\nPython"
lines = text.split('\n')
duplicates = set([line for line in lines if lines.count(line) > 1])
print(duplicates)

# 29. Validate strong passwords (min 8 chars, digit, uppercase)
pwds = ["Password1", "weak", "StrongPass2"]
strong = [p for p in pwds if re.fullmatch(r'(?=.*[A-Z])(?=.*\d).{8,}', p)]
print(strong)

# 30. Extract HTML <a> tag href values
html = '<a href="https://example.com">Link</a>'
hrefs = re.findall(r'<a\s+href="([^"]+)"', html)
print(hrefs)

# 31. Find all hexadecimal numbers in a file
text = "Colors: 0xFF, 0x1A3B"
hex_nums = re.findall(r'0x[0-9A-Fa-f]+', text)
print(hex_nums)

# 32. Extract content between brackets [...]
text = "List: [item1, item2, item3]"
content = re.findall(r'\[(.*?)\]', text)
print(content)

# 33. Validate MAC addresses
macs = ["00:1A:2B:3C:4D:5E", "001A2B3C4D5E"]
valid_macs = [m for m in macs if re.fullmatch(r'(?:[0-9A-Fa-f]{2}:){5}[0-9A-Fa-f]{2}', m)]
print(valid_macs)

# 34. Split a string at capital letters
s = "CamelCaseWord"
split_words = re.findall(r'[A-Z][a-z]*', s)
print(split_words)

# 35. Convert snake_case to camelCase in a string
s = "this_is_snake_case"
camel = re.sub(r'_(\w)', lambda m: m.group(1).upper(), s)
print(camel)
--------------------
Category 2: File Handling Exercises (35)

import os
import shutil
import csv
import json
from datetime import datetime, timedelta

# 36. Read a text file and count the number of lines
with open('example.txt') as f:
    lines = f.readlines()
print(len(lines))

# 37. Count the number of words in a file
with open('example.txt') as f:
    text = f.read()
word_count = len(text.split())
print(word_count)

# 38. Count character frequency in a file
from collections import Counter
char_count = Counter(text)
print(char_count)

# 39. Copy contents of one file to another
shutil.copyfile('example.txt', 'copy.txt')

# 40. Merge multiple text files into one
files = ['file1.txt', 'file2.txt']
with open('merged.txt', 'w') as outfile:
    for fname in files:
        with open(fname) as f:
            outfile.write(f.read() + '\n')

# 41. Reverse the content of a file line by line
with open('example.txt') as f:
    lines = f.readlines()
with open('reversed.txt', 'w') as f:
    f.writelines(lines[::-1])

# 42. Remove empty lines from a file
with open('example.txt') as f:
    non_empty = [line for line in f if line.strip()]
with open('no_empty.txt', 'w') as f:
    f.writelines(non_empty)

# 43. Append data to a CSV file
with open('data.csv', 'a', newline='') as f:
    writer = csv.writer(f)
    writer.writerow(['Alice', 30])

# 44. Read CSV and calculate average values of a column
with open('data.csv') as f:
    reader = csv.DictReader(f)
    values = [int(row['age']) for row in reader]
average = sum(values)/len(values)
print(average)

# 45. Convert a CSV file to JSON
with open('data.csv') as f:
    reader = csv.DictReader(f)
    data = list(reader)
with open('data.json', 'w') as f:
    json.dump(data, f, indent=2)

# 46. Read a log file and extract error lines
with open('log.txt') as f:
    errors = [line for line in f if 'ERROR' in line]
print(errors)

# 47. Split a large file into smaller files
chunk_size = 5
with open('bigfile.txt') as f:
    lines = f.readlines()
for i in range(0, len(lines), chunk_size):
    with open(f'part_{i//chunk_size+1}.txt', 'w') as f:
        f.writelines(lines[i:i+chunk_size])

# 48. Combine multiple CSV files into one
csv_files = ['a.csv', 'b.csv']
combined = []
for file in csv_files:
    with open(file) as f:
        combined.extend(f.readlines())
with open('combined.csv', 'w') as f:
    f.writelines(combined)

# 49. Detect duplicate lines in a file
with open('example.txt') as f:
    lines = f.read().splitlines()
duplicates = set([l for l in lines if lines.count(l) > 1])
print(duplicates)

# 50. Rename multiple files in a folder automatically
for i, filename in enumerate(os.listdir('folder')):
    ext = os.path.splitext(filename)[1]
    os.rename(os.path.join('folder', filename), os.path.join('folder', f'file_{i+1}{ext}'))

# 51. Delete files older than a given date
cutoff = datetime.now() - timedelta(days=30)
for f in os.listdir('folder'):
    path = os.path.join('folder', f)
    if os.path.isfile(path) and datetime.fromtimestamp(os.path.getmtime(path)) < cutoff:
        os.remove(path)

# 52. Count the number of words in all files in a folder
total_words = 0
for f in os.listdir('folder'):
    with open(os.path.join('folder', f)) as file:
        total_words += len(file.read().split())
print(total_words)

# 53. Replace all tabs with spaces in a file
with open('example.txt') as f:
    text = f.read()
text = text.replace('\t', '    ')
with open('example.txt', 'w') as f:
    f.write(text)

# 54. Remove lines containing a specific word
with open('example.txt') as f:
    lines = [line for line in f if 'delete' not in line]
with open('example.txt', 'w') as f:
    f.writelines(lines)

# 55. Find the longest line in a file
with open('example.txt') as f:
    longest = max(f, key=len)
print(longest)

# 56. Create a backup of all files in a folder
backup_dir = 'backup'
os.makedirs(backup_dir, exist_ok=True)
for f in os.listdir('folder'):
    shutil.copy(os.path.join('folder', f), backup_dir)

# 57. Extract email addresses from a text file
with open('example.txt') as f:
    text = f.read()
emails = re.findall(r'\b[\w.-]+@[\w.-]+\.\w+\b', text)
print(emails)

# 58. Count the number of occurrences of a specific word
word = 'Python'
with open('example.txt') as f:
    text = f.read()
count = text.count(word)
print(count)

# 59. Convert all text in a file to uppercase
with open('example.txt') as f:
    text = f.read()
with open('example.txt', 'w') as f:
    f.write(text.upper())

# 60. Read a log file and summarize occurrences of each log level
levels = {}
with open('log.txt') as f:
    for line in f:
        for lvl in ['INFO','WARN','ERROR']:
            if lvl in line:
                levels[lvl] = levels.get(lvl,0)+1
print(levels)

# 61. Detect the encoding of a file (requires chardet)
# pip install chardet
import chardet
with open('example.txt', 'rb') as f:
    result = chardet.detect(f.read())
print(result)

# 62. Check if a file exists before reading
if os.path.exists('example.txt'):
    with open('example.txt') as f:
        print(f.read())

# 63. Move all files of a certain type to a different folder
os.makedirs('txt_files', exist_ok=True)
for f in os.listdir('folder'):
    if f.endswith('.txt'):
        shutil.move(os.path.join('folder', f), 'txt_files')

# 64. Replace a word with another in all files of a folder
for f in os.listdir('folder'):
    path = os.path.join('folder', f)
    with open(path) as file:
        text = file.read()
    text = text.replace('old', 'new')
    with open(path, 'w') as file:
        file.write(text)

# 65. Read JSON files and extract specific keys
with open('data.json') as f:
    data = json.load(f)
values = [d['name'] for d in data]
print(values)

# 66. Convert multiple JSON files to CSV
json_files = ['a.json', 'b.json']
with open('output.csv','w',newline='') as csvfile:
    writer = None
    for jf in json_files:
        with open(jf) as f:
            data = json.load(f)
            if writer is None:
                writer = csv.DictWriter(csvfile, fieldnames=data[0].keys())
                writer.writeheader()
            writer.writerows(data)

# 67. Find all empty files in a directory
empty_files = [f for f in os.listdir('folder') if os.path.getsize(os.path.join('folder', f))==0]
print(empty_files)

# 68. Read a configuration file and update a value
import configparser
config = configparser.ConfigParser()
config.read('config.ini')
config['DEFAULT']['path'] = '/new/path'
with open('config.ini', 'w') as f:
    config.write(f)

# 69. Merge text files alphabetically by filename
files = sorted(os.listdir('folder'))
with open('merged.txt','w') as outfile:
    for fname in files:
        with open(os.path.join('folder', fname)) as f:
            outfile.write(f.read() + '\n')

# 70. Generate a word frequency report for a folder of text files
from collections import Counter
counter = Counter()
for f in os.listdir('folder'):
    with open(os.path.join('folder', f)) as file:
        counter.update(file.read().split())
print(counter.most_common(10))
--------------------
Category 3: Parsing Exercises (30)

import csv
import json
import xml.etree.ElementTree as ET
import re
from openpyxl import load_workbook

# 71. Parse a CSV file without using csv module
with open('data.csv') as f:
    data = [line.strip().split(',') for line in f]
print(data)

# 72. Extract structured data from unstructured logs
log = "ERROR 2026-03-07 User failed login"
match = re.search(r'(ERROR|INFO|WARN)\s+(\d{4}-\d{2}-\d{2})\s+(.*)', log)
if match:
    level, date, msg = match.groups()
    print(level, date, msg)

# 73. Parse JSON strings and access nested values
json_str = '{"user": {"name": "Alice", "age": 30}}'
data = json.loads(json_str)
print(data['user']['name'])

# 74. Parse XML files and extract elements
tree = ET.parse('example.xml')
root = tree.getroot()
for elem in root.findall('item'):
    print(elem.find('name').text, elem.find('price').text)

# 75. Extract data from HTML tables
html = """
<table>
<tr><th>Name</th><th>Age</th></tr>
<tr><td>Alice</td><td>30</td></tr>
<tr><td>Bob</td><td>25</td></tr>
</table>
"""
rows = re.findall(r'<tr>(.*?)</tr>', html, re.DOTALL)
for row in rows[1:]:
    cols = re.findall(r'<td>(.*?)</td>', row)
    print(cols)

# 76. Convert a Markdown table to CSV
md = "| Name | Age |\n|---|---|\n|Alice|30|\n|Bob|25|"
lines = md.splitlines()[2:]
with open('table.csv','w',newline='') as f:
    writer = csv.writer(f)
    for line in lines:
        writer.writerow([x.strip() for x in line.split('|')[1:-1]])

# 77. Parse Apache or Nginx log files
log = '127.0.0.1 - - [07/Mar/2026:12:30:00] "GET /index.html HTTP/1.1" 200 1024'
parts = re.match(r'(\S+) \S+ \S+ \[(.*?)\] "(.*?)" (\d+) (\d+)', log).groups()
print(parts)

# 78. Extract IP addresses and request paths from server logs
ips = re.findall(r'(\d{1,3}(?:\.\d{1,3}){3}) .*?"GET (.*?) HTTP', log)
print(ips)

# 79. Parse system configuration files (like .ini)
import configparser
config = configparser.ConfigParser()
config.read('config.ini')
print(config['DEFAULT']['path'])

# 80. Extract key-value pairs from a text document
text = "name: Alice\nage: 30\ncity: NY"
data = dict(re.findall(r'(\w+):\s*(.*)', text))
print(data)

# 81. Parse dates from multiple formats in a text
from datetime import datetime
dates = ["07-03-2026", "2026/03/07"]
parsed = []
for d in dates:
    try: parsed.append(datetime.strptime(d, "%d-%m-%Y"))
    except: parsed.append(datetime.strptime(d, "%Y/%m/%d"))
print(parsed)

# 82. Parse and validate JSON files before loading
with open('data.json') as f:
    try:
        data = json.load(f)
        print("Valid JSON")
    except json.JSONDecodeError:
        print("Invalid JSON")

# 83. Extract specific columns from a TSV file
with open('data.tsv') as f:
    cols = [line.strip().split('\t')[1] for line in f]  # column 2
print(cols)

# 84. Parse Git logs to get commit messages
git_log = """commit abc123
Author: Alice
Date: 2026-03-07

    Fixed bug"""
messages = re.findall(r'\n\n\s+(.*)', git_log)
print(messages)

# 85. Extract all links from an HTML page
html = '<a href="https://example.com">Link</a>'
links = re.findall(r'href="([^"]+)"', html)
print(links)

# 86. Parse Python requirements.txt and generate a list of packages
reqs = "numpy==1.21\npandas==1.3\n"
packages = [line.split('==')[0] for line in reqs.splitlines()]
print(packages)

# 87. Parse emails for subject, sender, and date
email = """From: alice@example.com
Date: Sat, 7 Mar 2026 12:30:00
Subject: Hello"""
sender = re.search(r'From: (.*)', email).group(1)
date = re.search(r'Date: (.*)', email).group(1)
subject = re.search(r'Subject: (.*)', email).group(1)
print(sender, date, subject)

# 88. Extract all table rows from an HTML string
html = "<tr><td>A</td><td>1</td></tr>"
rows = re.findall(r'<tr>(.*?)</tr>', html)
for r in rows:
    cols = re.findall(r'<td>(.*?)</td>', r)
    print(cols)

# 89. Parse SQL dump and extract table names
sql = "CREATE TABLE users(id INT); CREATE TABLE orders(id INT);"
tables = re.findall(r'CREATE TABLE (\w+)', sql)
print(tables)

# 90. Extract phone numbers from a text document
text = "Call me at 123-456-7890 or 987-654-3210"
phones = re.findall(r'\d{3}-\d{3}-\d{4}', text)
print(phones)

# 91. Parse server response headers
headers = """Content-Type: text/html
Content-Length: 1024"""
header_dict = dict(re.findall(r'([^:]+):\s*(.*)', headers))
print(header_dict)

# 92. Extract text from PDF files (requires PyPDF2)
# pip install PyPDF2
from PyPDF2 import PdfReader
reader = PdfReader('example.pdf')
text = ""
for page in reader.pages:
    text += page.extract_text()
print(text[:100])

# 93. Parse Excel files without using Pandas (with openpyxl)
wb = load_workbook('example.xlsx')
sheet = wb.active
for row in sheet.iter_rows(values_only=True):
    print(row)

# 94. Extract nested values from deeply nested dictionaries
data = {'a': {'b': {'c': 10}}}
value = data['a']['b']['c']
print(value)

# 95. Parse .log files and summarize error types
with open('log.txt') as f:
    errors = [line for line in f if 'ERROR' in line]
print(len(errors), "errors found")

# 96. Extract structured data from CSV with inconsistent delimiters
lines = ["name;age", "Alice;30", "Bob,25"]
data = [re.split(r'[;,]', line) for line in lines]
print(data)

# 97. Parse URLs to extract query parameters
url = "https://example.com/?id=123&name=Alice"
params = dict(re.findall(r'[?&](\w+)=(\w+)', url))
print(params)

# 98. Extract variables and values from a .env file
env = "DB_USER=root\nDB_PASS=1234"
env_dict = dict(re.findall(r'(\w+)=(.*)', env))
print(env_dict)

# 99. Parse HTML forms and extract input field names
html = '<form><input name="username"><input name="password"></form>'
fields = re.findall(r'name="([^"]+)"', html)
print(fields)

# 100. Parse Markdown headers to build a table of contents
md = "# Title\n## Section 1\n### Subsection"
headers = re.findall(r'^(#{1,6})\s*(.*)', md, re.MULTILINE)
toc = [(len(h[0]), h[1]) for h in headers]
print(toc)
--------------------
Category 4: Web Scraping Exercises (40)

# Install required packages if not already: pip install requests beautifulsoup4 selenium pandas

import requests
from bs4 import BeautifulSoup
import pandas as pd

# 101. Fetch a webpage using requests
url = "https://example.com"
response = requests.get(url)
html = response.text
print(html[:200])  # first 200 chars

# 102. Parse HTML with BeautifulSoup
soup = BeautifulSoup(html, 'html.parser')
print(soup.title.string)

# 103. Extract all links (<a> href)
links = [a['href'] for a in soup.find_all('a', href=True)]
print(links[:5])

# 104. Extract all images (<img> src)
images = [img['src'] for img in soup.find_all('img', src=True)]
print(images[:5])

# 105. Extract all paragraph texts
paragraphs = [p.get_text() for p in soup.find_all('p')]
print(paragraphs[:5])

# 106. Extract text from a specific class
texts = [div.get_text() for div in soup.find_all('div', class_='content')]
print(texts[:5])

# 107. Extract text from a specific ID
text = soup.find(id='main').get_text() if soup.find(id='main') else ''
print(text)

# 108. Extract all headings (h1-h6)
headings = [h.get_text() for h in soup.find_all(re.compile('^h[1-6]$'))]
print(headings)

# 109. Extract links containing a specific keyword
keyword_links = [a['href'] for a in soup.find_all('a', href=True) if 'blog' in a['href']]
print(keyword_links)

# 110. Extract all tables as DataFrames
tables = pd.read_html(html)
for table in tables:
    print(table.head())

# 111. Extract table by ID
table = soup.find('table', id='data-table')
rows = [[td.get_text() for td in tr.find_all(['td','th'])] for tr in table.find_all('tr')]
print(rows[:5])

# 112. Extract list items
items = [li.get_text() for li in soup.find_all('li')]
print(items[:5])

# 113. Extract meta tags
meta = {m.get('name', m.get('property')): m.get('content') for m in soup.find_all('meta') if m.get('content')}
print(meta)

# 114. Extract all scripts with src attribute
scripts = [s['src'] for s in soup.find_all('script', src=True)]
print(scripts)

# 115. Extract all stylesheets
css = [link['href'] for link in soup.find_all('link', rel='stylesheet')]
print(css)

# 116. Extract links ending with .pdf
pdf_links = [a['href'] for a in soup.find_all('a', href=True) if a['href'].endswith('.pdf')]
print(pdf_links)

# 117. Extract form input names
inputs = [i['name'] for i in soup.find_all('input', name=True)]
print(inputs)

# 118. Extract option values from select dropdowns
options = [o['value'] for o in soup.find_all('option', value=True)]
print(options)

# 119. Extract images with alt text
images_alt = [(img['src'], img.get('alt')) for img in soup.find_all('img', src=True)]
print(images_alt[:5])

# 120. Extract links and text together
link_texts = [(a.get_text(), a['href']) for a in soup.find_all('a', href=True)]
print(link_texts[:5])

# 121. Scrape multiple pages using pagination
base_url = "https://example.com/page/"
for i in range(1, 4):
    r = requests.get(base_url + str(i))
    soup = BeautifulSoup(r.text, 'html.parser')
    print(soup.title.string)

# 122. Extract tables and convert to CSV
for i, table in enumerate(pd.read_html(html)):
    table.to_csv(f'table_{i}.csv', index=False)

# 123. Extract all links and save to a file
with open('links.txt','w') as f:
    for link in links:
        f.write(link + '\n')

# 124. Extract headings and subheadings
headings = [h.get_text() for h in soup.find_all(['h1','h2'])]
print(headings)

# 125. Extract bold and italic text
bold_text = [b.get_text() for b in soup.find_all('b')]
italic_text = [i.get_text() for i in soup.find_all('i')]
print(bold_text[:5], italic_text[:5])

# 126. Scrape tables with specific class
table = soup.find('table', class_='data')
rows = [[td.get_text() for td in tr.find_all(['td','th'])] for tr in table.find_all('tr')]
print(rows[:5])

# 127. Extract links from a specific section
section = soup.find('div', class_='sidebar')
section_links = [a['href'] for a in section.find_all('a', href=True)]
print(section_links)

# 128. Extract article titles and summaries
articles = soup.find_all('article')
for a in articles:
    title = a.find('h2').get_text() if a.find('h2') else ''
    summary = a.find('p').get_text() if a.find('p') else ''
    print(title, summary)

# 129. Scrape images and download them
import os
os.makedirs('images', exist_ok=True)
for img in soup.find_all('img', src=True):
    img_url = img['src']
    r = requests.get(img_url)
    with open(os.path.join('images', os.path.basename(img_url)), 'wb') as f:
        f.write(r.content)

# 130. Extract links matching regex
pattern_links = [a['href'] for a in soup.find_all('a', href=True) if re.search(r'/products/\d+', a['href'])]
print(pattern_links)

# 131. Extract author names from articles
authors = [a.get_text() for a in soup.find_all('span', class_='author')]
print(authors)

# 132. Scrape nested elements
nested = soup.find_all('div', class_='post')
for div in nested:
    title = div.find('h3').get_text() if div.find('h3') else ''
    date = div.find('span', class_='date').get_text() if div.find('span', class_='date') else ''
    print(title, date)

# 133. Extract all text inside a tag and strip whitespace
texts = [t.get_text(strip=True) for t in soup.find_all('p')]
print(texts[:5])

# 134. Extract image URLs with specific extension
jpg_images = [img['src'] for img in soup.find_all('img', src=True) if img['src'].endswith('.jpg')]
print(jpg_images)

# 135. Scrape table data and convert to dictionary
table = soup.find('table')
data_dict = {}
for tr in table.find_all('tr'):
    cols = tr.find_all(['td','th'])
    if len(cols) >= 2:
        data_dict[cols[0].get_text()] = cols[1].get_text()
print(data_dict)

# 136. Extract links and check if they are internal or external
internal = [a['href'] for a in soup.find_all('a', href=True) if a['href'].startswith('/')]
external = [a['href'] for a in soup.find_all('a', href=True) if a['href'].startswith('http')]
print(internal[:5], external[:5])

# 137. Extract text from multiple pages (pagination)
for page in range(1,4):
    r = requests.get(f'https://example.com/page/{page}')
    soup = BeautifulSoup(r.text,'html.parser')
    print([h.get_text() for h in soup.find_all('h2')])

# 138. Extract data from JSON embedded in HTML
script = soup.find('script', type='application/ld+json')
if script:
    data = json.loads(script.string)
    print(data)

# 139. Extract links from a table
table = soup.find('table')
links_in_table = [a['href'] for a in table.find_all('a', href=True)]
print(links_in_table)

# 140. Scrape dynamic content using Selenium
# from selenium import webdriver
# driver = webdriver.Chrome()
# driver.get("https://example.com")
# content = driver.find_element("tag name", "body").text
# print(content[:200])
# driver.quit()
--------------------